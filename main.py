import openpyxl
import urllib.request
import xml.etree.ElementTree as ET
import json
import csv
import pandas as pd
# from openpyxl import Workbook

wb = openpyxl.load_workbook('2022 YIR Mailing List.xlsx', data_only=True)
print(wb.sheetnames)
sheet = wb['2022 combined list']
n_rows = 5

# Dynamically create dictionary from the spreadsheet
address_dict = {}
# Set to iterate over the number of columns
n_col = 0
# Remove max_row once ready for production
for col in sheet.iter_cols(min_row=1, max_row=n_rows, min_col=1):
    # Set row_n to be the id number for each row
    row_n = 0
    temp_dict = {}
    for row in sheet.iter_rows(min_row=1, max_row=n_rows, min_col=1):
        if row_n == 0:
            row_head = row[n_col].value
        else:
            temp_dict[row_n] = row[n_col].value
        row_n += 1
    n_col += 1
    temp_dict = {row_head: temp_dict}
    address_dict.update(temp_dict)
    # print(temp_dict)
print(address_dict)
# print(address_dict['First Name'][3])
list_dict = list(address_dict)
print(list_dict)

for k in address_dict['Addr1']:
    print(k, address_dict['Addr1'][k], address_dict['Addr2'][k], address_dict['City'][k])
#     x = address_dict.get('Addr1', {}).get('jesus', 'nothing')
# print(x)

# new_addr1_dict, new_addr2_dict, new_city_dict, new_state_dict, new_zip_dict, new_extzip_dict = {}
temp_addr1_dict, temp_addr2_dict, temp_city_dict, temp_state_dict\
    , temp_zip_dict, temp_extzip_dict, temp_err_dict = {}, {}, {}, {}, {}, {}, {}

for k in address_dict['Addr1']:
    requestXML = """
    <?xml version="1.0"?>
    <AddressValidateRequest USERID="292BENWA3717">
        <Revision>1</Revision>
        <Address ID="0">
            <Address1>"""+str(address_dict.get('Addr1', {}).get(k, '') or '')+"""</Address1>
            <Address2>"""+str(address_dict.get('Addr2', {}).get(k, '') or '')+"""</Address2>
            <City>"""+str(address_dict.get('City', {}).get(k, '') or '')+"""</City>
            <State>"""+str(address_dict.get('State', {}).get(k, '') or '')+"""</State>
            <Zip5>"""+str(address_dict.get('Zip', {}).get(k, '') or '')+"""</Zip5>
            <Zip4/>
        </Address>
    </AddressValidateRequest>
    """
    # print("XML\n", requestXML)

    #prepare xml string doc for query string
    docString = requestXML
    docString = docString.replace('\n','').replace('\t','')
    docString = urllib.parse.quote_plus(docString)

    url = "http://production.shippingapis.com/ShippingAPI.dll?API=Verify&XML=" + docString
    # print(url + "\n\n")

    response = urllib.request.urlopen(url)
    if response.getcode() != 200:
        print("Error making HTTP call:")
        print(response.info())
        exit()

    contents = response.read()
    # print(contents)

    root = ET.fromstring(contents)
    # print(root.text)

    for address in root.findall('Address'):
        if not address.find("Error"):
            # print("Address1: " + address.find("Address1").text)
            # print("Address2: " + address.find("Address2").text)
            # print("City:	 " + address.find("City").text)
            # print("State:	" + address.find("State").text)
            # print("Zip5:	 " + address.find("Zip5").text)
            # print("Zip4:	 " + address.find("Zip4").text)
            temp_addr1_dict[k] = address.find("Address1").text
            temp_addr2_dict[k] = address.find("Address2").text
            temp_city_dict[k] = address.find("City").text
            temp_state_dict[k] = address.find("State").text
            temp_zip_dict[k] = address.find("Zip5").text
            temp_extzip_dict[k] = address.find("Zip4").text
            temp_err_dict[k] = ""

        else:
            temp_addr1_dict[k] = ""
            temp_addr2_dict[k] = ""
            temp_city_dict[k] = ""
            temp_state_dict[k] = ""
            temp_zip_dict[k] = ""
            temp_extzip_dict[k] = ""
            for err in root.findall('Address/Error'):
                temp_err_dict[k] = err.find("Description").text
                # print("Error: ", err.find("Description").text)
        #     print("get error: ", root.text)

new_addr1_dict = {"New_Addr1": temp_addr1_dict}
new_addr2_dict = {"New_Addr2": temp_addr2_dict}
new_city_dict = {"New_City": temp_city_dict}
new_state_dict = {"New_State": temp_state_dict}
new_zip_dict = {"New_Zip": temp_zip_dict}
new_extzip_dict = {"New_ExtZip": temp_extzip_dict}
new_err_dict = {"Error_Description" : temp_err_dict}

address_dict = address_dict | new_addr1_dict | new_addr2_dict | new_city_dict\
               | new_state_dict | new_zip_dict | new_extzip_dict | new_err_dict
list_dict = list(address_dict)
# print(list_dict)
address_json = json.dumps(address_dict, indent=4)
print(address_json)

df = pd.read_json(address_json)

df.to_csv('clean_addresses.csv', encoding='utf-8', index=False)
